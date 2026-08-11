"""Export applicant records from Google Drive to an Excel workbook."""

import io
import json
import os
import sys
import tempfile
from datetime import datetime

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import phonenumbers
import pycountry

SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/drive.file",
]
CREDENTIALS_FILE = os.path.join(os.path.dirname(__file__), "credentials.json")
TOKEN_FILE = os.path.join(os.path.dirname(__file__), "token.json")
ROOT_FOLDER_ID = "1_y9GM6OtGX8mKrE4ROTCqtjunP-o9KEC"
NOTIFICATIONS_FOLDER_ID = "1RWPN91qoMQX7wP73LeXlPYp8LP1qLVby"
FOLDER_MIME_TYPE = "application/vnd.google-apps.folder"
IMAGE_EXTENSIONS = {".heic", ".heif", ".jpeg", ".jpg", ".png"}
COUNTRY_NAME_ALIASES = {
    "bolivia": "BO",
    "brunei": "BN",
    "cape verde": "CV",
    "czech republic": "CZ",
    "iran": "IR",
    "ivory coast": "CI",
    "kosovo": "XK",
    "laos": "LA",
    "moldova": "MD",
    "north korea": "KP",
    "palestine": "PS",
    "russia": "RU",
    "south korea": "KR",
    "syria": "SY",
    "tanzania": "TZ",
    "venezuela": "VE",
    "vietnam": "VN",
}
RECORD_FILENAMES = {
    "customer-information.json",
    "application-data.json",
    "customer-info.txt",
}
LEGACY_FIELD_NAMES = {
    "firstname": "firstName",
    "lastname": "lastName",
    "middlename": "middleName",
    "email": "email",
    "phone": "phone",
    "phonewhatsapp": "phone",
    "whatsapp": "phone",
    "countryofbirth": "countryOfBirth",
    "paymentstatus": "paymentStatus",
    "package": "package",
    "packagelabel": "packageLabel",
    "submittedat": "submittedAt",
    "submissiondate": "submittedAt",
    "submissionid": "submissionId",
}
HEADERS = [
    "First name",
    "Last name",
    "Middle name",
    "Email",
    "Phone / WhatsApp",
    "Country code",
    "Country of birth",
    "Payment status",
    "Files received",
    "Photo submitted",
    "Package",
    "Submitted at",
    "Submission ID",
]


def authenticate():
    if not os.path.exists(CREDENTIALS_FILE):
        sys.exit(
            "credentials.json not found. "
            "Download it from GCP → APIs & Services → Credentials and place it "
            "next to this script."
        )

    creds = None

    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)

    if not creds or not creds.valid or not creds.has_scopes(SCOPES):
        if creds and creds.expired and creds.refresh_token and creds.has_scopes(SCOPES):
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0, open_browser=False, prompt="consent")

        with open(TOKEN_FILE, "w") as fh:
            fh.write(creds.to_json())

    return creds


def list_files(service, query, fields):
    files = []
    page_token = None
    while True:
        response = service.files().list(
            q=query,
            spaces="drive",
            fields=f"nextPageToken, files({fields})",
            pageToken=page_token,
        ).execute()
        files.extend(response.get("files", []))
        page_token = response.get("nextPageToken")
        if not page_token:
            return files


def download_json(service, file_id):
    content = service.files().get_media(fileId=file_id).execute()
    return json.loads(content.decode("utf-8"))


def download_text(service, file_id):
    content = service.files().get_media(fileId=file_id).execute()
    return content.decode("utf-8")


def parse_legacy_text_record(content):
    record = {}
    for line in content.splitlines():
        if ":" not in line:
            continue
        label, value = line.split(":", 1)
        normalized_label = "".join(character for character in label.lower() if character.isalnum())
        field_name = LEGACY_FIELD_NAMES.get(normalized_label)
        if field_name:
            record[field_name] = value.strip()
    return record


def read_customer_record(service, children):
    record_file = next((item for item in children if item["name"] in RECORD_FILENAMES), None)
    if not record_file:
        return None

    try:
        if record_file["name"].endswith(".json"):
            return download_json(service, record_file["id"])
        return parse_legacy_text_record(download_text(service, record_file["id"]))
    except (json.JSONDecodeError, UnicodeDecodeError) as error:
        print(f"Skipping {record_file['name']}: invalid customer record ({error}).")
        return None


def applicants_from_record(record):
    applicants = record.get("applicants")
    if applicants:
        return applicants
    return [{
        "firstName": record.get("firstName", ""),
        "lastName": record.get("lastName", ""),
        "middleName": record.get("middleName", ""),
        "email": record.get("email", ""),
        "phone": record.get("phone", ""),
        "countryOfBirth": record.get("countryOfBirth", ""),
    }]


def country_dialing_code(country_name):
    if not country_name:
        return ""

    normalized_name = country_name.strip().lower()
    country_code = COUNTRY_NAME_ALIASES.get(normalized_name)
    if not country_code:
        try:
            country_code = pycountry.countries.lookup(country_name).alpha_2
        except LookupError:
            return ""

    dialing_code = phonenumbers.country_code_for_region(country_code)
    return f"+{dialing_code}" if dialing_code else ""


def format_submitted_at(timestamp):
    if not timestamp:
        return ""
    try:
        parsed = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
    except ValueError:
        return timestamp
    return parsed.strftime("%m/%d/%Y %H:%M")


def documents_for_submission(service, children):
    root_uploads = [
        item for item in children
        if item["mimeType"] != FOLDER_MIME_TYPE
        and item["name"] not in RECORD_FILENAMES
        and item["name"] != "upload-summary.txt"
    ]
    documents_folder = next(
        (item for item in children if item["name"] == "Documents" and item["mimeType"] == FOLDER_MIME_TYPE),
        None,
    )
    if not documents_folder:
        return root_uploads
    nested_uploads = list_files(
        service,
        f"'{documents_folder['id']}' in parents and trashed = false and mimeType != '{FOLDER_MIME_TYPE}'",
        "name",
    )
    return root_uploads + nested_uploads


def applicant_rows(service):
    submissions = list_files(
        service,
        f"'{ROOT_FOLDER_ID}' in parents and trashed = false and mimeType = '{FOLDER_MIME_TYPE}'",
        "id, name",
    )
    rows = []

    for submission in submissions:
        if submission["id"] == NOTIFICATIONS_FOLDER_ID:
            continue

        children = list_files(
            service,
            f"'{submission['id']}' in parents and trashed = false",
            "id, name, mimeType",
        )
        record = read_customer_record(service, children)
        if not record:
            print(f"Skipping {submission['name']}: no supported customer record found.")
            continue

        documents = documents_for_submission(service, children)
        file_names = [item["name"] for item in documents]
        has_photo = any(
            os.path.splitext(name.lower())[1] in IMAGE_EXTENSIONS for name in file_names
        )

        for applicant in applicants_from_record(record):
            rows.append([
                applicant.get("firstName", ""),
                applicant.get("lastName", ""),
                applicant.get("middleName", ""),
                applicant.get("email", ""),
                applicant.get("phone", ""),
                country_dialing_code(applicant.get("countryOfBirth", "")),
                applicant.get("countryOfBirth", ""),
                record.get("paymentStatus", ""),
                "True" if file_names else "False",
                "True" if has_photo else "False",
                record.get("packageLabel") or record.get("package", ""),
                format_submitted_at(record.get("submittedAt", "")),
                record.get("submissionId", ""),
            ])

    return rows


def create_workbook(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Applicants"
    sheet.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for row in rows:
        sheet.append(row)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
        sheet.column_dimensions[column[0].column_letter].width = width

    return workbook


def upload_workbook(service, workbook):
    timestamp = datetime.now().strftime("%m_%d_%Y_%H_%M")
    filename = f"{timestamp}_applicants.xlsx"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temporary_file:
        temporary_path = temporary_file.name

    try:
        workbook.save(temporary_path)
        metadata = {"name": filename, "parents": [ROOT_FOLDER_ID]}
        media = MediaFileUpload(
            temporary_path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        uploaded = service.files().create(body=metadata, media_body=media, fields="id, webViewLink").execute()
    finally:
        os.remove(temporary_path)

    return filename, uploaded


def main():
    print("Authenticating with Google Drive...")
    creds = authenticate()
    service = build("drive", "v3", credentials=creds)

    print("Scanning applicant submission folders...")
    rows = applicant_rows(service)
    workbook = create_workbook(rows)
    filename, uploaded = upload_workbook(service, workbook)

    print(f"Exported {len(rows)} applicant(s) to {filename}.")
    print(f"Drive file: {uploaded.get('webViewLink', uploaded['id'])}")


if __name__ == "__main__":
    main()
